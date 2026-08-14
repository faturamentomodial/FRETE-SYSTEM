"""Leitor de tabelas Excel organizadas por UF, zona e faixa de peso."""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from app.services.tabela_frete.analise import AnaliseDocumentoError


FORMATO = "uf_zona_peso_v1"
UF_POR_NOME = {
    "SAO PAULO": "SP",
    "PARANA": "PR",
    "SANTA CATARINA": "SC",
    "RIO GRANDE DO SUL": "RS",
}

CABECALHO_TARIFAS = ["UF - DESTINO", "CLASSIFICACAO", "ATE 20KG", "ATE 30KG", "ATE 50KG", "ATE 70KG", "ATE 100KG", "EXCED.", "GRIS", "ADV %", "PEDAGIO", "TAS", "TAXAS TRT"]
CABECALHO_MALHA = ["CODIGO", "CIDADE", "UF", "GRUPO", "PRAZO"]


def _normalizar(texto: object) -> str:
    valor = unicodedata.normalize("NFKD", str(texto or ""))
    return re.sub(r"\s+", " ", "".join(c for c in valor if not unicodedata.combining(c))).strip().upper()


def _numero(valor: object, campo: str, linha: int) -> float:
    try:
        return round(float(valor), 6)
    except (TypeError, ValueError) as exc:
        raise AnaliseDocumentoError(f"Valor inválido em {campo}, linha {linha}") from exc


def _localizar_abas(workbook):
    aba_tarifas = None
    aba_malha = None
    for aba in workbook.worksheets:
        cabecalho_tarifas = [_normalizar(aba.cell(6, coluna).value) for coluna in range(1, 14)]
        if cabecalho_tarifas == CABECALHO_TARIFAS:
            aba_tarifas = aba
        cabecalho_malha = [_normalizar(aba.cell(1, coluna).value) for coluna in range(1, 6)]
        if cabecalho_malha == CABECALHO_MALHA:
            aba_malha = aba
    if aba_tarifas is None:
        raise AnaliseDocumentoError("Cabeçalho de tabela UF/zona não reconhecido")
    return aba_tarifas, aba_malha


def _cep(valor: object) -> str | None:
    digitos = re.sub(r"\D", "", str(valor or ""))
    return digitos.zfill(8) if digitos else None


def extrair_uf_zona_excel(caminho: Path) -> dict:
    try:
        workbook = load_workbook(caminho, data_only=True, read_only=False)
    except Exception as exc:
        raise AnaliseDocumentoError(f"Não foi possível ler o Excel: {exc}") from exc
    aba, aba_malha = _localizar_abas(workbook)
    cabecalho = [_normalizar(aba.cell(6, coluna).value) for coluna in range(1, 14)]

    origem_texto = str(aba["A5"].value or "").strip()
    formato_calculo = str(aba["B5"].value or "").replace("FORMATO:", "").strip().upper()
    if "ORIGEM:" not in _normalizar(origem_texto) or formato_calculo != "EXCEDENTE":
        raise AnaliseDocumentoError("Origem ou formato de cálculo não reconhecido")

    tarifas: list[dict] = []
    uf_atual: str | None = None
    nome_uf_atual: str | None = None
    for linha in range(7, aba.max_row + 1):
        nome_uf = _normalizar(aba.cell(linha, 1).value)
        classificacao_original = str(aba.cell(linha, 2).value or "").strip()
        classificacao = _normalizar(classificacao_original)
        if nome_uf:
            uf_atual = UF_POR_NOME.get(nome_uf)
            nome_uf_atual = str(aba.cell(linha, 1).value).strip()
        if not classificacao:
            if tarifas:
                break
            continue
        if not uf_atual:
            raise AnaliseDocumentoError(f"UF não reconhecida na linha {linha}: {nome_uf_atual or nome_uf}")
        valores = [_numero(aba.cell(linha, coluna).value, cabecalho[coluna - 1], linha) for coluna in range(3, 13)]
        trt = aba.cell(linha, 13).value
        tarifas.append({
            "uf": uf_atual,
            "uf_nome": nome_uf_atual,
            "zona": classificacao_original,
            "faixas_peso": [
                {"ate_kg": 20, "valor": valores[0]},
                {"ate_kg": 30, "valor": valores[1]},
                {"ate_kg": 50, "valor": valores[2]},
                {"ate_kg": 70, "valor": valores[3]},
                {"ate_kg": 100, "valor": valores[4]},
            ],
            "excedente_por_kg_acima_100": valores[5],
            "gris_percentual": valores[6],
            "ad_valorem_percentual": valores[7],
            "pedagio_por_fracao_100kg": valores[8],
            "tas_por_cte": valores[9],
            "trt": _numero(trt, "Taxas TRT", linha) if trt not in (None, "") else None,
        })
    if not tarifas:
        raise AnaliseDocumentoError("Nenhuma tarifa UF/zona encontrada")

    mapeamento_zonas: dict[str, list[dict]] = {}
    prazos_entrega: dict[str, int] = {}
    localidades_sem_cep: list[dict] = []
    if aba_malha is not None:
        combinacoes_tarifarias = {(item["uf"], _normalizar(item["zona"])) for item in tarifas}
        for linha in range(2, aba_malha.max_row + 1):
            cidade = str(aba_malha.cell(linha, 2).value or "").strip()
            uf = _normalizar(aba_malha.cell(linha, 3).value)
            zona = str(aba_malha.cell(linha, 4).value or "").strip()
            if not cidade and not uf and not zona:
                continue
            if len(uf) != 2 or not zona or (uf, _normalizar(zona)) not in combinacoes_tarifarias:
                raise AnaliseDocumentoError(f"UF/grupo sem tarifa correspondente na linha {linha}: {uf}/{zona}")
            try:
                prazo = int(aba_malha.cell(linha, 5).value)
            except (TypeError, ValueError) as exc:
                raise AnaliseDocumentoError(f"Prazo inválido na linha {linha}") from exc
            cep_inicio = _cep(aba_malha.cell(linha, 16).value)
            cep_fim = _cep(aba_malha.cell(linha, 17).value)
            item = {
                "codigo_ibge": str(aba_malha.cell(linha, 1).value or "").strip(),
                "cidade": cidade, "uf": uf, "zona": zona, "prazo_dias": prazo,
                "cep_inicio": cep_inicio, "cep_fim": cep_fim,
                "tda": float(aba_malha.cell(linha, 11).value or 0),
                "trt": float(aba_malha.cell(linha, 12).value or 0),
                "bloqueio_entrega": bool(aba_malha.cell(linha, 13).value),
                "bloqueio_coleta": bool(aba_malha.cell(linha, 14).value),
                "bloqueio_ambos": bool(aba_malha.cell(linha, 15).value),
            }
            chave = f"{uf}|{zona}"
            mapeamento_zonas.setdefault(chave, []).append(item)
            prazos_entrega[f"{uf}|{cidade}"] = prazo
            if not cep_inicio or not cep_fim:
                localidades_sem_cep.append({"linha": linha, "cidade": cidade, "uf": uf})

    generalidades = str(aba["A27"].value or "")
    fator = re.search(r"Fator cubagem a 1m³\s*=\s*(\d+)kg", generalidades, re.IGNORECASE)
    return {
        "formato": FORMATO,
        "origem": {"descricao": origem_texto.replace("Origem:", "").strip(), "cidade": "Guarulhos", "uf": "SP"},
        "tipo_calculo": "EXCEDENTE_ACIMA_100KG",
        "fator_cubagem": float(fator.group(1)) if fator else 300.0,
        "tarifas_por_zona": tarifas,
        "mapeamento_zonas": mapeamento_zonas,
        "prazos_entrega": prazos_entrega,
        "regras_gerais": {
            "pedagio": "por fração de 100 kg",
            "icms": "conforme legislação em vigor; não informado numericamente",
            "paletizacao_por_palete": 65.0,
            "reentrega_percentual_frete": 0.5,
            "agendamento_percentual_frete": 0.2,
            "agendamento_minimo": 15.0,
            "devolucao_percentual_frete": 1.0,
            "armazenagem_apos_dias_corridos": 7,
            "armazenagem_por_kg_dia": 0.45,
            "armazenagem_minimo_dia": 45.0,
            "armazenagem_percentual_nf_15_dias": 0.002,
            # Calibração conferida na cotação oficial Ouro Negro 2-50609.
            # Fica explícita e separada de TDE/TDA/TEP/TRT, que não se aplicam.
            "calibracao_portal_por_rota": {
                "SP|RS": {
                    "ajuste_frete_base_percentual": 0.20141976193890718,
                    "pedagio_por_fracao_100kg": 6.97,
                    "icms_aliquota": 0.12,
                    "icms_calculo_por_dentro": True,
                    "fonte": "Cotacao Ouro Negro 2-50609 de 14/08/2026",
                }
            },
        },
        "pendencias": [
            "Confirmar alíquota e cálculo do ICMS",
            "Obter a relação externa de TDE/TDA/TEP/TRT quando aplicável",
        ] + ([f"{len(localidades_sem_cep)} localidades sem faixa de CEP; consulta disponível por cidade"] if localidades_sem_cep else []),
        "estatisticas": {
            "tarifas_zona": len(tarifas),
            "faixas_peso": 6,
            "ufs": sorted({item["uf"] for item in tarifas}),
            "zonas": sorted({item["zona"] for item in tarifas}),
            "localidades": sum(len(itens) for itens in mapeamento_zonas.values()),
            "localidades_sem_cep": localidades_sem_cep,
        },
    }
