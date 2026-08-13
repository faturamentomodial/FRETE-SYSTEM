"""Leitor do modelo comercial RTE Rodonaves KM/peso com malha por CEP."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from app.services.tabela_frete.analise import AnaliseDocumentoError


FORMATO = "rodonaves_km_peso_v1"


def _cep(valor) -> int:
    digitos = re.sub(r"\D", "", str(valor or ""))
    if not digitos:
        raise AnaliseDocumentoError("CEP vazio na malha de cobertura")
    return int(digitos.zfill(8))


def _faixa_km(texto: str) -> tuple[int, int | None]:
    numeros = [int(item.replace(".", "")) for item in re.findall(r"\d[\d.]*", texto)]
    if "ACIMA" in texto.upper():
        return numeros[0] + 1, None
    if len(numeros) < 2:
        raise AnaliseDocumentoError(f"Faixa de KM inválida: {texto}")
    return numeros[0], numeros[1]


def extrair_rodonaves_excel(caminho: Path) -> dict:
    try:
        workbook = load_workbook(caminho, data_only=True, read_only=True)
    except Exception as exc:
        raise AnaliseDocumentoError(f"Não foi possível ler o Excel: {exc}") from exc
    obrigatorias = {"TABELA", "ANEXO I ", "CEP'S ZONA RESTRIÇÃO - RJ", "CEP'S ZONA DE RISCO - SPO", "CEP'S CENTRO EXPANDIDO - SPO"}
    faltantes = obrigatorias - set(workbook.sheetnames)
    if faltantes:
        raise AnaliseDocumentoError(f"Abas obrigatórias ausentes: {', '.join(sorted(faltantes))}")

    tabela = workbook["TABELA"]
    razao_cliente = str(tabela["B6"].value or "").replace("Razão Social:", "").strip()
    documento_cliente = str(tabela["J6"].value or "").strip()
    origem_texto = str(tabela["C12"].value or "SÃO PAULO - SP")
    matriz = []
    for linha in range(15, 39):
        descricao = str(tabela.cell(linha, 2).value or "").strip()
        km_min, km_max = _faixa_km(descricao)
        valores = [float(tabela.cell(linha, coluna).value) for coluna in range(4, 10)]
        matriz.append({
            "descricao": descricao, "km_min": km_min, "km_max": km_max,
            "ate_10": valores[0], "ate_20": valores[1], "ate_40": valores[2],
            "ate_60": valores[3], "ate_100": valores[4], "por_kg_acima_100": valores[5],
        })

    anexo = workbook["ANEXO I "]
    coberturas = []
    empresas = set()
    for valores in anexo.iter_rows(min_row=2, values_only=True):
        if not valores[5] or not valores[6] or valores[9] is None or not valores[14] or not valores[15]:
            continue
        empresas.add(str(valores[17] or "").strip())
        coberturas.append({
            "cidade": str(valores[5]).strip(), "uf": str(valores[6]).strip().upper(),
            "cep_inicio": _cep(valores[14]), "cep_fim": _cep(valores[15]),
            "km": int(valores[9]), "prazo_pj": int(valores[11]), "prazo_pf": int(valores[12]),
            "frequencia": str(valores[10] or "").strip(), "polo": str(valores[16] or "").strip(),
            "empresa_malha": str(valores[17] or "").strip(),
            "taxa_despacho": round(float(valores[21] or 0), 6),
            "taxa_cidade": round(float(valores[22] or 0), 6),
            "taxa_emex": round(float(valores[23] or 0), 6),
            "taxa_final": round(float(valores[24] or 0), 6),
        })
    if len(coberturas) < 100:
        raise AnaliseDocumentoError("A malha de CEP parece incompleta")

    def faixas_taxa(nome_aba: str, linha_inicial: int, indice_inicio: int, indice_fim: int, indice_valor: int) -> list[dict]:
        aba = workbook[nome_aba]
        itens = []
        for valores in aba.iter_rows(min_row=linha_inicial, values_only=True):
            if not valores[indice_inicio] or not valores[indice_fim]:
                continue
            try:
                itens.append({
                    "cep_inicio": _cep(valores[indice_inicio]), "cep_fim": _cep(valores[indice_fim]),
                    "valor": round(float(valores[indice_valor] or 0), 6),
                })
            except (ValueError, TypeError, AnaliseDocumentoError):
                continue
        return itens

    restricao_rj = faixas_taxa("CEP'S ZONA RESTRIÇÃO - RJ", 4, 0, 2, 4)
    risco_sp = faixas_taxa("CEP'S ZONA DE RISCO - SPO", 3, 1, 2, 7)
    centro_sp = []
    for valores in workbook["CEP'S CENTRO EXPANDIDO - SPO"].iter_rows(min_row=2, values_only=True):
        if valores[0]:
            try:
                centro_sp.append(_cep(valores[0]))
            except AnaliseDocumentoError:
                pass

    return {
        "formato": FORMATO,
        "cliente": {"razao_social": razao_cliente, "cnpj_cpf": documento_cliente},
        "transportadora": {"nome": "RTE Rodonaves", "razao_social": "RODONAVES TRANSPORTES E ENCOMENDAS LTDA"},
        "origem": origem_texto,
        "fator_cubagem": 300.0,
        "peso_limite_kg": 7000.0,
        "matriz_tarifas": matriz,
        "coberturas": coberturas,
        "zonas": {"restricao_rj": restricao_rj, "risco_sp": risco_sp, "centro_expandido_sp": centro_sp},
        "regras": {
            "frete_valor_percentual": float(tabela["K46"].value), "frete_valor_minimo": float(tabela["K47"].value),
            "gris_sul_sudeste_percentual_ate_10k": float(tabela["K56"].value),
            "gris_sul_sudeste_percentual_acima_10k": float(tabela["K57"].value),
            "gris_sul_sudeste_minimo": float(tabela["K55"].value),
            "gris_demais_percentual": float(tabela["K61"].value), "gris_demais_minimo": float(tabela["K60"].value),
            "pedagio_por_100kg": float(tabela["K64"].value),
            "taxa_administrativa": {
                "valor": float(tabela["K51"].value),
                "ufs": ["GO", "MG", "DF", "MT", "MS", "RO", "AC", "PA", "RR", "AP", "TO", "AM"],
            },
            "taxas_entrega_por_peso": [
                {
                    "cidades": ["ITAJAI", "PORTO ALEGRE", "CURITIBA", "NAVEGANTES"],
                    "peso_limite": 60, "ate_limite": float(tabela["K77"].value),
                    "acima_limite": float(tabela["K78"].value),
                },
                {
                    "cidades": ["CAMBORIU", "BALNEARIO CAMBORIU"],
                    "peso_limite": 60, "ate_limite": float(tabela["K80"].value),
                    "acima_limite": float(tabela["K81"].value),
                },
                {
                    "cidades": ["GOIANIA", "APARECIDA DE GOIANIA", "SENADOR CANEDO", "ABADIA DE GOIAS", "GOIANIRA", "TRINDADE", "INHUMAS"],
                    "peso_limite": 100, "ate_limite": float(tabela["K83"].value),
                    "acima_limite": float(tabela["K84"].value),
                },
            ],
            "icms": {
                "calculo_por_dentro": True,
                "aliquota_reduzida": 0.07,
                "aliquota_padrao": 0.12,
                "origens_reduzida": ["SP", "MG", "RJ", "PR", "SC", "RS"],
                "destinos_reduzida": ["AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "PA", "PB", "PE", "PI", "RN", "RO", "RR", "SE", "TO"],
            },
            "centro_sp_faixas": [
                {"peso_max": 100, "valor": float(tabela["D96"].value)}, {"peso_max": 200, "valor": float(tabela["E96"].value)},
                {"peso_max": 300, "valor": float(tabela["F96"].value)}, {"peso_max": 500, "valor": float(tabela["G96"].value)},
                {"peso_max": 1000, "valor": float(tabela["H96"].value)}, {"peso_max": 1500, "valor": float(tabela["I96"].value)},
            ],
            "impostos_inclusos": False, "prazo_utilizado": "PJ_DIAS_UTEIS",
        },
        "estatisticas": {
            "faixas_km": len(matriz), "coberturas_cep": len(coberturas), "restricoes_rj": len(restricao_rj),
            "riscos_sp": len(risco_sp), "ceps_centro_sp": len(centro_sp), "empresas_malha": sorted(item for item in empresas if item),
        },
    }
