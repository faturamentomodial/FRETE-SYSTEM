from pathlib import Path

from openpyxl import Workbook

from app.services.tabela_frete.calculo_uf_zona import calcular_uf_zona
from app.services.tabela_frete.uf_zona_excel import CABECALHO_MALHA, CABECALHO_TARIFAS, extrair_uf_zona_excel


def _arquivo_duas_abas(caminho: Path) -> None:
    wb = Workbook()
    tarifas = wb.active
    tarifas.title = "Precos"
    tarifas["A5"] = "Origem: Guarulhos/SP"
    tarifas["B5"] = "FORMATO: EXCEDENTE"
    for coluna, valor in enumerate(CABECALHO_TARIFAS, 1):
        tarifas.cell(6, coluna, valor)
    valores = ["Parana", "Interior", 10, 20, 30, 40, 50, 1.5, .0015, .0015, 6.29, 5.6, None]
    for coluna, valor in enumerate(valores, 1):
        tarifas.cell(7, coluna, valor)
    malha = wb.create_sheet("Malha")
    cabecalho = CABECALHO_MALHA + ["SEG", "TER", "QUA", "QUI", "SEX", "TDA", "TRT", "BLOQ ENT", "BLOQ COL", "BLOQ AMBOS", "CEP INICIAL", "CEP FINAL"]
    for coluna, valor in enumerate(cabecalho, 1):
        malha.cell(1, coluna, valor)
    linha = ["4100103", "ABATIA", "PR", "Interior", 5, None, "S", None, None, "S", 10, 0, None, None, None, 86460000, 86464999]
    for coluna, valor in enumerate(linha, 1):
        malha.cell(2, coluna, valor)
    wb.save(caminho)


def test_extrai_tarifas_malha_e_prazos_de_abas_distintas(tmp_path):
    caminho = tmp_path / "tabela.xlsx"
    _arquivo_duas_abas(caminho)

    dados = extrair_uf_zona_excel(caminho)

    assert dados["estatisticas"]["localidades"] == 1
    assert dados["mapeamento_zonas"]["PR|Interior"][0]["cep_inicio"] == "86460000"
    assert dados["prazos_entrega"]["PR|ABATIA"] == 5


def test_calcula_tabela_uf_zona_importada(tmp_path):
    caminho = tmp_path / "tabela.xlsx"
    _arquivo_duas_abas(caminho)
    dados = extrair_uf_zona_excel(caminho)

    resultado = calcular_uf_zona(dados, {"peso": 25, "valor_nf": 1000, "destino_uf": "PR", "destino_cep": "86460-000"})

    assert resultado["status"] == "success"
    assert resultado["frete_base"] == 20
    assert resultado["prazo_dias"] == 5
    assert resultado["valor_total"] == 44.89


def test_reproduz_cotacao_oficial_ouro_negro_2_50609():
    dados = {
        "fator_cubagem": 300,
        "tarifas_por_zona": [{
            "uf": "RS", "zona": "Interior II",
            "faixas_peso": [{"ate_kg": 20, "valor": 50.597}, {"ate_kg": 30, "valor": 53.3995}, {"ate_kg": 50, "valor": 55.784}],
            "excedente_por_kg_acima_100": .9405, "gris_percentual": .0015,
            "ad_valorem_percentual": .0015, "pedagio_por_fracao_100kg": 6.29, "tas_por_cte": 5.6, "trt": None,
        }],
        "mapeamento_zonas": {"RS|Interior II": [{
            "cidade": "SANTA CRUZ DO SUL", "uf": "RS", "zona": "Interior II", "prazo_dias": 6,
            "cep_inicio": "96800000", "cep_fim": "96874999", "tda": 0, "trt": 0,
            "bloqueio_entrega": False, "bloqueio_ambos": False,
        }]},
        "regras_gerais": {"calibracao_portal_por_rota": {"SP|RS": {
            "ajuste_frete_base_percentual": .20141976193890718, "pedagio_por_fracao_100kg": 6.97,
            "icms_aliquota": .12, "icms_calculo_por_dentro": True,
        }}},
    }

    resultado = calcular_uf_zona(dados, {
        "peso": 20, "valor_nf": 5290, "origem_uf": "SP", "destino_uf": "RS",
        "destino_cep": "96810-062", "volume_total_m3": 0.12, "quantidade_volumes": 2,
    })

    assert resultado["peso_cubado_kg"] == 36
    assert resultado["prazo_dias"] == 6
    assert next(item for item in resultado["taxas_detalhadas"] if item["tipo"] == "PEDAGIO")["valor"] == 6.97
    assert next(item for item in resultado["taxas_detalhadas"] if item["tipo"] == "ICMS")["valor"] == 13.02
    assert resultado["valor_total"] == 108.48
